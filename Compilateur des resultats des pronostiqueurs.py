# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import os
from openpyxl import load_workbook
from openpyxl import Workbook

def copy_sheet(source_file, destination_file):
    # Chargement du classeur source
    source_wb = load_workbook(source_file, data_only=True)  # Utiliser data_only=True pour récupérer les valeurs des cellules
    
    # Parcours de toutes les feuilles du classeur source
    for sheet_name in source_wb.sheetnames:
        if sheet_name == "Feuille pronostiqueur":
            # Sélection de la feuille à copier
            source_ws = source_wb[sheet_name]
            
            # Création d'une nouvelle feuille dans le classeur de destination
            new_sheet_name = f"{sheet_name} {len(destination_file.sheetnames)}"
            destination_ws = destination_file.create_sheet(new_sheet_name)
            
            # Copie des valeurs de la plage de cellules A1:P27
            for row in source_ws.iter_rows(min_row=1, max_row=27, min_col=1, max_col=16):
                destination_ws.append([cell.value for cell in row])
    
    # Sauvegarde du classeur de destination
    destination_file.save("Synthese des resultats des pronostiqueurs.xlsx")

def main():
    # Dossier contenant les fichiers sources
    source_folder = '//fichiers01/usersW10$/Paul FAVIER/Documents/Experiences-PF/Experiences-PF/Paris tournoi babyfoot/Resultats des pronostiqueurs'
    
    # Chemin vers le fichier Excel de destination
    destination_file_path = '//fichiers01/usersW10$/Paul FAVIER/Documents/Experiences-PF/Experiences-PF/Paris tournoi babyfoot/Excel pronostiqueur.xlsx'
    
    # Chargement du classeur de destination
    if os.path.exists(destination_file_path):
        destination_wb = load_workbook(destination_file_path)
    else:
        destination_wb = Workbook()
        destination_wb.save(destination_file_path)
    
    # Parcourir tous les fichiers dans le dossier source
    for filename in os.listdir(source_folder):
        if filename.endswith('.xlsx'):
            source_file = os.path.join(source_folder, filename)
            # Copier la feuille "Feuille pronostiqueur" dans le fichier de destination
            copy_sheet(source_file, destination_wb)

    # Fermeture du classeur de destination
    destination_wb.close()

if __name__ == "__main__":
    main()
